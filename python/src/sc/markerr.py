"""
Mark errors in the results file
"""
import openpyxl
import re
import os
import math
import numpy as np
import pandas as pd
from mtl import verbalnumbers

import mtl.verbalnumbers.hebrew as hebnum

lexical_classes = hebnum.ones, hebnum.tens, hebnum.hundreds, hebnum.thousands


# noinspection PyMethodMayBeStatic
class ErrorAnalyzer(object):

    #------------------------------------------------------
    def __init__(self, digit_mapping=None, subj_id_transformer=None, consider_thousand_as_digit=True, accuracy_per_digit=False,
                 fail_on_segment_order_error=False, subj_id_in_xls=True, in_col_names=None, phonological_error_flds=(),
                 set_per_subject=None):
        """

        :param phonological_error_flds: List of xls columns which contain number of phonological errors. All these columns will be summed.
        :param digit_mapping:
        :param subj_id_transformer: Function for transfoming the subject ID field
        :param consider_thousand_as_digit:  Whether the word "thousand" should count towards digit errors also in numbers with 5 or 6 digits
        :param accuracy_per_digit: This concerns the output files per word/morpheme: whether to compute the word/digit accuracy
                for each specific word or less precisely. Value=True is impossible if the target contains duplicate digits.
        """
        self._digit_mapping = {str(d): d for d in range(0, 10)}
        if digit_mapping is not None:
            self._digit_mapping.update(digit_mapping)

        self.subj_id_transformer = subj_id_transformer
        self.consider_thousand_as_digit = consider_thousand_as_digit
        self.accuracy_per_digit = accuracy_per_digit
        self.fail_on_segment_order_error = fail_on_segment_order_error
        self.phonological_error_flds = phonological_error_flds
        self.subj_id_in_xls = subj_id_in_xls

        self.in_col_names = dict(block='Block', condition='Condition', itemnum='ItemNum', target='target', response='response',
                                 nwords='NWordsPerTarget', exclude='exclude', manual='manual')
        if in_col_names is not None:
            for k, v in in_col_names.items():
                assert k in self.in_col_names, 'Invalid in_col_names - unexpected column "{}"'.format(k)
                self.in_col_names[k] = v

        optional_col_keys = 'exclude', 'manual'

        self.xls_mandatory_cols = tuple(c for k, c in self.in_col_names.items() if c is not None and k not in optional_col_keys)
        self.xls_optional_cols = tuple(c for k, c in self.in_col_names.items() if c is not None and k in optional_col_keys)
        self.xls_cols = (('Subject',) if subj_id_in_xls else ()) + self.xls_mandatory_cols
        self.xls_out_cols = ('Subject',)+self.xls_mandatory_cols+self.xls_optional_cols+ \
                            ('NTargetDigits', 'NMissingWords', 'PMissingWords', 'NMissingDigits',
                             'PMissingDigits', 'NMissingClasses', 'PMissingClasses', 'PMissingMorphemes')

        if len(self.phonological_error_flds) > 0:
            self.xls_out_cols += 'NPhonologicalErrors',

        if set_per_subject is None:
            self.fixed_value_per_subject = {}
        else:
            self.fixed_value_per_subject = { sid: {} for sid in set_per_subject['subjid'] }
            for i, sid in enumerate(set_per_subject['subjid']):
                self.fixed_value_per_subject[sid] = {cn: set_per_subject[cn][i] for cn in set_per_subject.keys() if cn != 'subjid'}
            self.xls_out_cols += tuple(cn for cn in set_per_subject.keys() if cn != 'subjid')


    #------------------------------------------------------
    def run_for_worksheet(self, in_fn, worksheet='data', out_dir=None, out_fn_prefix='data_coded'):
        """
        Analyze the error rates (digit, class, morpheme, word) in each trial

        :param in_fn: Excel file with the raw data (uncoded)
        :param worksheet: Name of worksheet to read
        :param out_dir: Directory for output files
        :param out_fn_prefix:
        """
        self.run_for_worksheets(in_fn, [worksheet], out_dir, out_fn_prefix)


    #------------------------------------------------------
    def run_for_worksheets(self, in_fn, worksheets=None, out_dir=None, out_fn_prefix='data_coded'):
        """
        Analyze the error rates (digit, class, morpheme, word) in each trial

        :param set_per_subject: values to set for each participant. This is a dict with a 'subjid' entry and
                                one additional entry for each column to set
        """

        out_wb, out_ws = self.create_output_workbook()
        wb = openpyxl.load_workbook(in_fn)
        if worksheets is None:
            worksheets = [ws.title for ws in wb.worksheets]

        result_per_word = []
        n_excluded = []
        n_phonerr = []

        ok = True
        out_row_num = 2

        for worksheet in worksheets:
            print('\nProcessing worksheet [{}]...'.format(worksheet))

            try:
                in_ws, col_inds = self._open_worksheet(wb, worksheet, in_fn)
            except ValueError as e:
                print('>>> ERROR (worksheet ignored): {}'.format(e))
                continue

            found_empty_rows = False
            subj_n_excluded = 0
            subj_n_phonerr = 0

            if 'manual' in self.in_col_names:
                nrep = sum('repeat' in str(in_ws.cell(rownum, col_inds[self.in_col_names['manual']]).value) for rownum in range(2, in_ws.max_row+1))
                if nrep > 0:
                    print('{}: {} excluded&repeated trials'.format(worksheet.title(), nrep))

            for rownum in range(2, in_ws.max_row+1):
                if self.fixed_value_per_subject is not None and worksheet in self.fixed_value_per_subject:
                    self.set_fixed_values(self.fixed_value_per_subject[worksheet], out_ws, out_row_num)

                rc = self.parse_row(in_ws, out_ws, rownum, out_row_num, col_inds, result_per_word, worksheet)

                if rc == 'empty':
                    found_empty_rows = True
                    continue

                elif rc == 'excluded':
                    subj_n_excluded += 1

                elif rc == 'error':
                    ok = False

                else:
                    subj_n_phonerr += rc
                    out_row_num += 1

                    if found_empty_rows:
                        print('ERROR: row {} in worksheet "{}" contains data but there were few empty rows previously. Skipped.'.
                              format(rownum, worksheet))
                        ok = False

            n_excluded.append(subj_n_excluded)
            n_phonerr.append(int(subj_n_phonerr))

            #-- Last row is not needed
            out_ws.delete_rows(out_row_num)

        if ok:
            print('{} rows were processed, no errors found.'.format(out_row_num-1))
        else:
            print('Some errors were encountered.')

        out_ws.freeze_panes = out_ws['A2']
        self.auto_col_width(out_ws)

        if out_dir is None:
            print('WARNING: results were not saved for {}'.format(in_fn))

        else:
            out_wb.save(out_dir + os.sep + out_fn_prefix + '.xlsx')
            pd.DataFrame(result_per_word).to_csv(out_dir + os.sep + out_fn_prefix + '_words.csv', index=False)

            subjstat = pd.DataFrame(dict(subject=worksheets, n_excluded=n_excluded))
            if len(self.phonological_error_flds) > 0:
                subjstat['n_phonerr'] = n_phonerr
            subjstat.to_csv(out_dir + os.sep + out_fn_prefix + '_subjstat.csv', index=False)


    #------------------------------------------------------
    def parse_row(self, in_ws, out_ws, rownum, out_rownum, col_inds, result_per_word, worksheet):
        """
        Parse a single row, copy it to the output file
        Return: the number of phonological errors; or a string reflecting the error/issue
        """

        if self.in_col_names['exclude'] in col_inds:
            exclude = in_ws.cell(rownum, col_inds[self.in_col_names['exclude']]).value
            if exclude == 1:
                return 'excluded'

        if self.subj_id_in_xls:
            subj_id = in_ws.cell(rownum, col_inds[self.in_col_names['subject']]).value
        else:
            subj_id = worksheet

        raw_target = in_ws.cell(rownum, col_inds[self.in_col_names['target']]).value
        raw_response = in_ws.cell(rownum, col_inds[self.in_col_names['response']]).value

        if self.in_col_names['nwords'] is None:
            n_target_words = None
        else:
            n_target_words = in_ws.cell(rownum, col_inds[self.in_col_names['nwords']]).value

            if n_target_words is None:
                print("Error in line {}: 'NWordsPerTarget' was not specified".format(rownum))
                return 'error'

        if raw_target is None:
            return 'empty'

        #-- Save basic columns

        subj_id = subj_id if self.subj_id_transformer is None else self.subj_id_transformer(subj_id)
        self._save_value(out_ws, out_rownum, 'Subject', subj_id)

        copy_cols = self.xls_mandatory_cols + tuple([c for c in self.xls_optional_cols if self.in_col_names[c] in col_inds])
        for colname in copy_cols:
            self._save_value(out_ws, out_rownum, colname, in_ws.cell(rownum, col_inds[colname]).value)

        #-- Analyze target & response

        target, target_segments = self.parse_target(raw_target, rownum)

        if n_target_words is None:
            n_target_words = len(target)
        elif n_target_words != len(target):
            print("Error in line {}: Invalid number of words ({}), expecting {} words".format(rownum, len(target), n_target_words))
            return 'error'

        target_word_said, target_digit_said, n_class_errs = self.analyze_response(raw_response, target, target_segments, rownum)
        if target_word_said is None:
            return 'error'

        n_word_errs = sum([digsaid is False for digsaid in target_word_said])
        n_digit_errs = sum([digsaid is False for digsaid in target_digit_said])

        #-- Phonological errors
        if len(self.phonological_error_flds) > 0:
            n_phonerr = [in_ws.cell(rownum, col_inds[c]).value for c in self.phonological_error_flds]
            n_phonerr = [n for n in n_phonerr if not _isnull(n)]
            if sum(not isinstance(n, (int, float)) for n in n_phonerr) > 0:
                print('Error in line {}: invalid number of phonological errors ({})'.format(rownum, n_phonerr))
                n_phonerr = 0
            else:
                n_phonerr = sum(n_phonerr)
                self._save_value(out_ws, out_rownum, 'NPhonologicalErrors', n_phonerr)

        else:
            n_phonerr = 0

        #-- Save performance measures

        n_target_digits = sum([t.digit is not None for t in target])

        self._save_value(out_ws, out_rownum, 'NMissingWords', n_word_errs)
        self._save_value(out_ws, out_rownum, 'NMissingDigits', n_digit_errs)
        self._save_value(out_ws, out_rownum, 'NMissingClasses', n_class_errs)

        self._save_value(out_ws, out_rownum, 'NTargetDigits', n_target_digits)
        self._save_value(out_ws, out_rownum, 'PMissingWords', n_word_errs / n_target_words)
        self._save_value(out_ws, out_rownum, 'PMissingDigits', n_digit_errs / n_target_digits)
        self._save_value(out_ws, out_rownum, 'PMissingClasses', n_class_errs / n_target_words)
        self._save_value(out_ws, out_rownum, 'PMissingMorphemes', (n_class_errs + n_digit_errs) / (n_target_words + n_target_digits))

        self.custom_process_row(in_ws, out_ws, rownum, out_rownum, col_inds)

        self._save_accuracy_per_word(subj_id, in_ws, rownum, col_inds, target, target_word_said, target_digit_said, raw_response, raw_target,
                                     result_per_word)

        return n_phonerr


    #------------------------------------------------------------------------------
    def custom_process_row(self, in_ws, out_ws, in_rownum, out_rownum, col_inds):
        pass


    #------------------------------------------------------------------------------
    def parse_target(self, raw_target, rownum):
        target_segments = self.parse_target_or_response(raw_target, rownum)
        target = np.array(self.collapse_segments(target_segments)[::-1])  # put the ones word in position 1
        return target, target_segments


    #------------------------------------------------------------------------------
    def _save_accuracy_per_word(self, subj_id, in_ws, rownum, col_inds, target, target_word_said, target_digit_said,
                                raw_response, raw_target, result_per_word):

        block = None if self.in_col_names['block'] is None else in_ws.cell(rownum, col_inds[self.in_col_names['block']]).value
        cond_name = in_ws.cell(rownum, col_inds['Condition']).value
        item_num = in_ws.cell(rownum, col_inds['ItemNum']).value
        n_target_words = len(target_word_said)

        for i, (word_said, digit_said, target_word) in enumerate(zip(target_word_said, target_digit_said, target)):

            lexical_class_order = lexical_classes.index(target_word.lexical_class) if target_word.lexical_class in lexical_classes else ''

            r = dict(subject=subj_id,
                     block=block,
                     condition=cond_name,
                     item_num=item_num,
                     n_target_words=n_target_words,
                     target=raw_target,
                     response=raw_response,
                     word_order=n_target_words-i,
                     word_class=target_word.lexical_class,
                     word_class_order=lexical_class_order,
                     target_word=target_word,
                     word_ok=1 if word_said else 0,
                     digit_ok=None if digit_said is None else 1 if digit_said else 0,
                     )

            result_per_word.append(r)


    #------------------------------------------------------
    def analyze_response(self, raw_response, target, target_segments, rownum):
        """
        Analyze the target-response matching and save the results onto the Excel worksheet

        Returns:
        - a bool array with one entry per target word, indicating whether that word was said
        - a bool array with one entry per target word, indicating whether the word's digit was said (None if the target word has no digit)
        - The number of unsaid target classes

        :param raw_response: The response as a string
        :param target:
        :param target_segments:
        :param rownum:
        """

        response_segments = self.parse_response(raw_response, rownum, target_segments)
        if target_segments is None or response_segments is None:
            return [None] * 3

        response = self.collapse_segments(response_segments)

        target_word_said, target_digit_said = self.target_items_said(target, response)

        n_unsaid_target_classes = self._n_missing_classes(target, response)

        return target_word_said, target_digit_said, n_unsaid_target_classes


    #------------------------------------------------------
    def _save_value(self, out_ws, rownum, colname, value):
        out_ws.cell(rownum, 1 + self.xls_out_cols.index(colname)).value = value


    #------------------------------------------------------
    def _n_missing_classes(self, target, response):
        target = [t.lexical_class for t in target]
        response = [r.lexical_class for r in response]

        for r in response:
            if r in target:
                target.remove(r)

        return len(target)


    #------------------------------------------------------
    def target_items_said(self, target_words, response_words):
        """ Return an array of bool: for each target item, whether it was said or not """

        #---- Step 1: fully-correct words

        tmp_target_words = list(target_words)
        response_words = list(response_words)

        #-- Loop through response, mark each said word
        for i, resp in enumerate(response_words):
            if resp is None:
                continue

            try:
                tmp_target_words[tmp_target_words.index(resp)] = None
                response_words[i] = None
            except ValueError:
                #-- Exception from target_words.index() - i.e., response word was not said
                pass

        target_word_said = [t is None for t in tmp_target_words]

        #---- Step 2: digits said with incorrect class

        target_digit_said = list(target_word_said)
        for i, t in enumerate(target_words):
            if t.digit is None:
                target_digit_said[i] = None

        remaining_target_digit_inds = {t.digit: i for i, t in enumerate(tmp_target_words) if t is not None and t.digit is not None}
        remaining_response_digits = [r.digit for r in response_words if r is not None and r.digit is not None]

        #-- Loop through response, mark each said digit
        for i, resp in enumerate(remaining_response_digits):
            if resp in remaining_target_digit_inds:
                trg_ind = remaining_target_digit_inds[resp]
                del remaining_target_digit_inds[resp]
                target_digit_said[trg_ind] = True
                tmp_target_words[trg_ind] = None

        return target_word_said, target_digit_said


    #------------------------------------------------------
    def n_pairs(self, target, response, tind):
        """
        Find adjacent words with (class1, class2), and check if they appeared as an adjacent pair in the response

        Return:
        n_pairs_found - the number of pairs that existed in the target and the response
        n_mismatching_order - the number of pairs that existed in the target, and in the response existed with position violation (= not adjacent,
                              or not in the same order).

        :param target: list of words
        :param response: list of words
        :param tind: The index of the relevant target word
        """

        word1 = target[tind]
        word2 = target[tind+1]

        if word1 not in response or word2 not in response:
            return 0, 0

        correct_order_inds = [i for i, (r1, r2) in enumerate(zip(response[:-1], response[1:])) if r1 == word1 and r2 == word2]

        n_mismatching_order = int(len(correct_order_inds) == 0)

        return 1, n_mismatching_order


    #------------------------------------------------------
    def parse_response(self, response_str, rownum, target_segments):

        if response_str is None:
            return None

        if response_str in ('+', '!', 'v', 'V'):
            return target_segments

        if response_str == '-':
            return []

        response_str = str(response_str)

        #-- Check if there are optional things
        m = re.match('(.*);(.+)', response_str)
        if m is None:
            response_segments_unknown_loc = []
        else:
            response_str = m.group(1)
            response_segments_unknown_loc = self.parse_target_or_response(m.group(2), rownum)
            if response_segments_unknown_loc is None:
                return None

        try:
            response_segments = self.parse_target_or_response(response_str, rownum)
            if response_segments is None:
                return None
        except ValueError as e:
            print('Error in line {} (line ignored): {} '.format(rownum, e))
            return None

        target_has_duplicate_segments = len(target_segments) != len(set(target_segments))

        if len(response_segments) != len(target_segments) and ['correct'] in response_segments:
            print('WARNING: "+" is ambiguous because the target and response have different number of segments. ' +
                  'Line {} ignored'.format(rownum))
            return None

        #-- swap "+" with the corresponding target value
        for i, r in enumerate(response_segments):
            if list(r) == ['correct']:
                #-- double validation -- in case we have order mismatch. Validate this only if the tar
                if i >= len(target_segments) or (not target_has_duplicate_segments and target_segments[i] in response_segments):
                    print('WARNING: "+" is ambiguous because the target and response have different order of segments. ' +
                          'Line {} ignored'.format(rownum))
                    if self.fail_on_segment_order_error:
                        return None

                response_segments[i] = target_segments[i]

        return response_segments + response_segments_unknown_loc


    #------------------------------------------------------
    def collapse_segments(self, parsed_segments):
        return [x for pn in parsed_segments for x in pn]


    #------------------------------------------------------
    def parse_target_or_response(self, raw_text, rownum):
        """
        Return a list of segments, each of which is a list of words
        """

        if isinstance(raw_text, float):
            raw_text = "{:.0f}".format(raw_text)
        elif isinstance(raw_text, int):
            raw_text = "{:}".format(raw_text)

        segments = [e.strip() for e in raw_text.split('/')]

        parsed_segments = []
        for seg in segments:
            m = re.match('^([0-9,]*)\\s*t\\s*([0-9,]+)?$', seg)

            if m is None:
                parsed_segment = [self.parse_segment_into_word_list(seg)]
            else:
                parsed_segment = self._parse_pre_thousand_segment(m, seg)
                if m.group(2) is not None:
                    parsed_segment.append(self.parse_segment_into_word_list(m.group(2)))

            if None in parsed_segment:  # invalid format
                print('WARNING: unsupported target/response format: "{}" -- line {} ignored'.format(raw_text, rownum))
                return None

            #-- combine parts of the parsed segment
            parsed_segment = [e for seg in parsed_segment for e in seg]

            parsed_segments.append(tuple(parsed_segment))

        return parsed_segments


    #------------------------------------------------------
    def _parse_pre_thousand_segment(self, matcher, segment):
        """ Parse the 'thousand' and the preceding digits """

        if len(matcher.group(1)) == 0:
            # -- The word "thousand" with no preceding digit
            return [self.parse_segment_into_word_list('t')]

        elif len(matcher.group(1)) == 1:
            # -- A 4-digit number: the "thousand" is combined with the preceding digit
            return [self.parse_segment_into_word_list(matcher.group(1) + '000')]

        elif len(matcher.group(1)) in (2, 3):
            # -- A 5- or 6-digit number: the "thousand" is a separate word
            return [self.parse_segment_into_word_list(matcher.group(1)), self.parse_segment_into_word_list('t')]

        else:
            raise Exception('Unsupported format: {}'.format(segment))


    #------------------------------------------------------
    def parse_segment_into_word_list(self, segment):
        """
        Parse a number into a series of words

        :param segment: a string describing one grammatical segment (one number)
        """

        if segment in ('+', '!'):   # "+" is correct; "!" is TBD
            return ['correct']

        if segment in ('-', '?'):   # "-" means omission; "?" means "I don't know"
            return []

        segment = segment.replace(',', '')  # delete commas

        result = verbalnumbers.hebrew.number_to_words(segment, digit_mapping=self._digit_mapping)

        if not self.consider_thousand_as_digit:
            one_thousand = verbalnumbers.general.NumberWord(verbalnumbers.hebrew.thousands, 1)
            decimal_word_thousand = verbalnumbers.general.NumberWord(verbalnumbers.hebrew.decword_thousand, None)
            for i, w in enumerate(result):
                if w == one_thousand:
                    result[i] = decimal_word_thousand

        return result


    #------------------------------------------------------
    def set_fixed_values(self, fixed_values, out_ws, out_row_num):
        for k, v in fixed_values.items():
            self._save_value(out_ws, out_row_num, k, v)


    #------------------------------------------------------
    def _open_worksheet(self, wb, worksheet, filename):
        if len(wb.worksheets) == 1:
            ws = wb.worksheets[0]
        else:
            sheet_names = [s.title for s in wb.worksheets]
            if worksheet not in sheet_names:
                raise ValueError('{} does not contain any worksheet named "{}"'.format(filename, worksheet))
            ws = wb.get_sheet_by_name(worksheet)

        col_inds = self._xls_structure(ws)

        return ws, col_inds


    def _xls_structure(self, ws):
        result = {}
        for i in range(1, ws.max_column+1):
            col_name = ws.cell(1, i).value
            result[col_name] = i

        missing_cols = [c for c in self.xls_cols if c not in result]
        if len(missing_cols) > 0:
            raise ValueError('Invalid file format: columns {} are missing'.format(",".join(missing_cols)))

        return result


    #------------------------------------------------------
    def create_output_workbook(self):

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        for i_col, colname in enumerate(self.xls_out_cols):
            ws.cell(1, i_col+1).value = colname

        return wb, ws


    #------------------------------------------------------
    def auto_col_width(self, ws):
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            ws.column_dimensions[col[0].column_letter].width = max_length


def _isnull(v):
    return v is None or v == '' or (isinstance(v, float) and math.isnan(v))

