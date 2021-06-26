
import openpyxl
import re
import pandas as pd

lexical_classes = 'unit', 'decade', 'hundred', 'unit', 'decade', 'hundred'

xls_copy_cols = 'Block', 'Condition', 'ItemNum', 'target', 'response', 'NWordsPerTarget'
xls_cols = ('Subject', ) + xls_copy_cols + ('NMissingWords', 'NMissingDigits', 'NMissingClasses')
xls_out_cols = ('Subject', ) + xls_copy_cols + \
               ('NTargetDigits', 'NMissingWords', 'PMissingWords', 'NMissingDigits',
                'PMissingDigits', 'NMissingClasses', 'PMissingClasses', 'PMissingMorphemes')

xls_optional_cols = 'manual', 'WordOrder'


#------------------------------------------------------
def analyze_errors(in_fn, out_dir, consider_thousand_as_digit, use_teens=False, subj_id_parser=None):

    in_ws, col_inds = _open_input_file(in_fn)
    out_wb, out_ws = create_output_workbook()

    result_per_word = dict(subject=[], block=[], condition=[], itemNum=[], target=[], response=[], nTargetWords=[],
                           wordOK=[], digitOK=[], classOK=[])

    result_per_morpheme = dict(subject=[], block=[], condition=[], itemNum=[], target=[], response=[],
                               nTargetWords=[], nTargetDigits=[], nTargetMorphemes=[],
                               morpheme_type=[], correct=[])

    ok = True
    for rownum in range(2, in_ws.max_row+1):
        ok = parse_row(in_ws, out_ws, rownum, col_inds, result_per_word,
                       result_per_morpheme, consider_thousand_as_digit, use_teens, subj_id_parser) and ok

    if not ok:
        print('Some errors were encountered. Set 1 in the "manual" column to override automatic error encoding.')

    out_ws.freeze_panes = out_ws['A2']
    auto_col_width(out_ws)
    out_wb.save(out_dir + '/data_coded.xlsx')
    pd.DataFrame(result_per_word).to_csv(out_dir + '/data_coded_words.csv', index=False)
    pd.DataFrame(result_per_morpheme).to_csv(out_dir + '/data_coded_morphemes.csv', index=False)


#------------------------------------------------------
def parse_row(in_ws, out_ws, rownum, col_inds, result_per_word, result_per_morpheme, consider_thousand_as_digit, use_teens,
              subj_id_parser):
    """
    Parse a single row, copy it to the output file
    Return True if processed OK.

    # If the value in the "manual" column is 1, don't do anything - just copy the error rates from the corresponding columns
    """

    subj_id = in_ws.cell(rownum, col_inds['Subject']).value
    target_str = in_ws.cell(rownum, col_inds['target']).value
    if subj_id is None and target_str is None:
        print('Warning: row #{} seems empty, ignored'.format(rownum))
        return True

    out_ws.cell(rownum, xls_out_cols.index('Subject')+1).value = subj_id if subj_id_parser is None else subj_id_parser(subj_id)

    for colname in xls_copy_cols:
        out_ws.cell(rownum, xls_out_cols.index(colname) + 1).value = in_ws.cell(rownum, col_inds[colname]).value

    n_target_words = in_ws.cell(rownum, col_inds['NWordsPerTarget']).value

    target_segments = parse_target_or_response(target_str, rownum, use_teens)
    target = collapse_segments(target_segments)
    if n_target_words != len(target):
        print("Error in line {}: Invalid number of words ({}), expecting {} words".format(rownum, len(target), n_target_words))
        return False

    if consider_thousand_as_digit:
        target_digits = [t[1] for t in target]
    else:
        target_digits = [t[1] for t in target if t[0] != 'thousand']

    n_target_digits = len(target_digits)

    if 'manual' in col_inds and in_ws.cell(rownum, col_inds['manual']).value in (1, '1'):

        n_word_errs = _nullto0(in_ws.cell(rownum, col_inds['NMissingWords']).value)
        if 'WordOrder' in col_inds:
            n_word_errs -= _nullto0(in_ws.cell(rownum, col_inds['WordOrder']).value)

        n_digit_errs = _nullto0(in_ws.cell(rownum, col_inds['NMissingDigits']).value)
        n_class_errs = _nullto0(in_ws.cell(rownum, col_inds['NMissingClasses']).value)

    else:

        response_segments = parse_response(in_ws.cell(rownum, col_inds['response']).value, rownum, target_segments, use_teens)
        if target_segments is None or response_segments is None:
            return False

        response = collapse_segments(response_segments)

        n_word_errs = n_missing_target_items(target, response)
        n_class_errs = _n_missing_classes(target, response)

        if consider_thousand_as_digit:
            response_digits = [r[1] for r in response]
        else:
            response_digits = [r[1] for r in response if r[0] != 'thousand']

        n_digit_errs = n_missing_target_items(target_digits, response_digits)

    _save_value(out_ws, rownum, 'NTargetDigits', n_target_digits)
    _save_value(out_ws, rownum, 'NMissingWords', n_word_errs)
    _save_value(out_ws, rownum, 'NMissingDigits', n_digit_errs)
    _save_value(out_ws, rownum, 'NMissingClasses', n_class_errs)
    _save_value(out_ws, rownum, 'PMissingWords', n_word_errs / n_target_words)
    _save_value(out_ws, rownum, 'PMissingDigits', n_digit_errs / n_target_digits)
    _save_value(out_ws, rownum, 'PMissingClasses', n_class_errs / n_target_words)
    _save_value(out_ws, rownum, 'PMissingMorphemes', (n_class_errs + n_digit_errs) / (n_target_words + n_target_digits))

    for i in range(n_target_words):
        result_per_word['subject'].append(subj_id)
        result_per_word['block'].append(in_ws.cell(rownum, col_inds['Block']).value)
        result_per_word['condition'].append(in_ws.cell(rownum, col_inds['Condition']).value)
        result_per_word['itemNum'].append(in_ws.cell(rownum, col_inds['ItemNum']).value)
        result_per_word['target'].append(target_str)
        result_per_word['response'].append(in_ws.cell(rownum, col_inds['response']).value)
        result_per_word['nTargetWords'].append(n_target_words)
        result_per_word['wordOK'].append(1 if i >= n_word_errs else 0)
        result_per_word['digitOK'].append(1 if i >= n_digit_errs else 0)
        result_per_word['classOK'].append(1 if i >= n_class_errs else 0)

        result_per_morpheme['subject'].append(subj_id)
        result_per_morpheme['block'].append(in_ws.cell(rownum, col_inds['Block']).value)
        result_per_morpheme['condition'].append(in_ws.cell(rownum, col_inds['Condition']).value)
        result_per_morpheme['itemNum'].append(in_ws.cell(rownum, col_inds['ItemNum']).value)
        result_per_morpheme['target'].append(target_str)
        result_per_morpheme['response'].append(in_ws.cell(rownum, col_inds['response']).value)
        result_per_morpheme['nTargetWords'].append(n_target_words)
        result_per_morpheme['nTargetDigits'].append(n_target_digits)
        result_per_morpheme['nTargetMorphemes'].append(n_target_words + n_target_digits)
        result_per_morpheme['morpheme_type'].append('class')
        result_per_morpheme['correct'].append(1 if i >= n_class_errs else 0)

    for i in range(n_target_digits):
        result_per_morpheme['subject'].append(subj_id)
        result_per_morpheme['block'].append(in_ws.cell(rownum, col_inds['Block']).value)
        result_per_morpheme['condition'].append(in_ws.cell(rownum, col_inds['Condition']).value)
        result_per_morpheme['itemNum'].append(in_ws.cell(rownum, col_inds['ItemNum']).value)
        result_per_morpheme['target'].append(target_str)
        result_per_morpheme['response'].append(in_ws.cell(rownum, col_inds['response']).value)
        result_per_morpheme['nTargetWords'].append(n_target_words)
        result_per_morpheme['nTargetDigits'].append(n_target_digits)
        result_per_morpheme['nTargetMorphemes'].append(n_target_words + n_target_digits)
        result_per_morpheme['morpheme_type'].append('digit')
        result_per_morpheme['correct'].append(1 if i >= n_digit_errs else 0)

    return True


#------------------------
def _nullto0(v):
    if v is None or v == '':
        return 0
    else:
        return v


#------------------------------------------------------
def _save_value(out_ws, rownum, colname, value):
    out_ws.cell(rownum, 1 + xls_out_cols.index(colname)).value = value


#------------------------------------------------------
def _n_missing_classes(target, response):
    target = [t[0] for t in target]
    response = [r[0] for r in response]

    for r in response:
        if r in target:
            target.remove(r)

    return len(target)


#------------------------------------------------------
def n_missing_target_items(target_items, response_items):
    # No. of target items that are not in the response

    tmp = list(target_items)
    for r in response_items:
        if r in tmp:
            tmp.remove(r)

    return len(tmp)


#------------------------------------------------------
def parse_response(response_str, rownum, target_segments, use_teens):

    if response_str is None:
        return None

    if response_str in ('+', '!'):
        return target_segments

    if response_str == '-':
        return []

    response_str = str(response_str)

    #-- Check if there are optional things
    m = re.match('(.*);(.+)', response_str)
    if m is None:
        response_segments_unknown_loc = []
    else:
        response_str = m.group(1)
        response_segments_unknown_loc = parse_target_or_response(m.group(2), rownum, use_teens)
        if response_segments_unknown_loc is None:
            return None

    response_segments = parse_target_or_response(response_str, rownum, use_teens)
    if response_segments is None:
        return None

    target_has_duplicate_segments = len(target_segments) != len(set(target_segments))

    if len(response_segments) != len(target_segments) and ['correct'] in response_segments:
        print('WARNING: "+" is ambiguous because the target and response have different number of segments. '+
              'Line {} ignored'.format(rownum))
        return None

    #-- swap "+" with the corresponding target value
    for i, r in enumerate(response_segments):
        if r == ['correct']:
            #-- double validation -- in case we have order mismatch. Validate this only if the tar
            if not target_has_duplicate_segments and target_segments[i] in response_segments:
                print('WARNING: "+" is ambiguous because the target and response have different order of segments. ' +
                      'Line {} ignored'.format(rownum))
                return None

            response_segments[i] = target_segments[i]

    return response_segments + response_segments_unknown_loc


#------------------------------------------------------
def collapse_segments(parsed_segments):
    return [x for pn in parsed_segments for x in pn]


#------------------------------------------------------
def parse_target_or_response(raw_text, rownum, use_teens):
    """
    Return a list of segments, each of which is a list of words

    :param raw_text:
    :param rownum:
    """

    if isinstance(raw_text, float):
        raw_text = "{:.0f}".format(raw_text)
    elif isinstance(raw_text, int):
        raw_text = "{:}".format(raw_text)

    segments = [e.strip() for e in raw_text.split('/')]

    parsed_segments = []
    for seg in segments:
        m = re.match('^([0-9,]*)\\s*t\\s*([0-9,]+)?$', seg)

        if m is None:
            parsed_segment = [ parse_segment_into_word_list(seg, use_teens) ]
        else:
            parsed_segment = _parse_pre_thousand_segment(m, seg, use_teens)
            if m.group(2) is not None:
                parsed_segment.append(parse_segment_into_word_list(m.group(2), use_teens))

        if None in parsed_segment:  # invalid format
            print('WARNING: unsupported target/response format: "{}" -- line {} ignored'.format(raw_text, rownum))
            return None

        #-- combine parts of the parsed segment
        parsed_segment = [e for seg in parsed_segment for e in seg]

        parsed_segments.append(tuple(parsed_segment))

    return parsed_segments


#------------------------------------------------------
def _parse_pre_thousand_segment(matcher, segment, use_teens):
    """ Parse the 'thousand' and the preceding digits """

    if len(matcher.group(1)) == 0:
        # -- The word "thousand" with no preceding digit
        return [ parse_segment_into_word_list('t', use_teens) ]

    elif len(matcher.group(1)) == 1:
        # -- A 4-digit number: the "thousand" is combined with the preceding digit
        return [ parse_segment_into_word_list(matcher.group(1) + '000', False) ]

    elif len(matcher.group(1)) in (2, 3):
        # -- A 5- or 6-digit number: the "thousand" is a separate word
        return [parse_segment_into_word_list(matcher.group(1), use_teens), parse_segment_into_word_list('t', False)]

    else:
        raise Exception('Unsupported format: {}'.format(segment))


#------------------------------------------------------
def parse_segment_into_word_list(segment, use_teens):
    """
    Parse a number into a series of words
    """

    if segment == 't' or segment == '1000':
        return [('thousand', 1)]

    if segment in ('+', '!'):
        return ['correct']

    if segment == '-':
        return []

    segment = segment.replace(',', '')  # delete commas

    if re.match('^\\d+$', segment) is None:
        return None

    digits = [int(d) for d in segment[::-1]]
    ndigits = len(segment)
    if use_teens and ndigits > 1 and digits[1] == 1 and digits[0] != 0:
        digits[1] = 0
        digits[0] += 10

    result = []
    for i, digit in enumerate(digits):
        if ndigits == 4 and i == 3:
            result.append(('thousand', digit))

        elif i == 0 and digit > 10:
            result.append(('teens', digit-10))

        elif digit != 0:
            result.append((lexical_classes[i], digit))

        if i == 2 and ndigits > 4:
            # todo the following works for this experiment, but may not work for others:
            #      the word "thousand" in 5/6-digit numbers may count as a digit error versus numbers like 31
            result.append(('thousand', 1))

    return result[::-1]


#------------------------------------------------------
def _open_input_file(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.get_sheet_by_name('data')

    col_inds = _xls_structure(ws)

    return ws, col_inds


def _xls_structure(ws):
    result = {}
    for i in range(1, ws.max_column+1):
        col_name = ws.cell(1, i).value
        if col_name in xls_cols + xls_optional_cols:
            result[col_name] = i

    missing_cols = [c for c in xls_cols if c not in result]
    if len(missing_cols) > 0:
        raise Exception('Invalid file format: columns {} are missing'.format(",".join(missing_cols)))

    return result


#------------------------------------------------------
def create_output_workbook():

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    for i_col, colname in enumerate(xls_out_cols):
        ws.cell(1, i_col+1).value = colname

    return wb, ws


#------------------------------------------------------
def auto_col_width(ws):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        ws.column_dimensions[col[0].column_letter].width = max_length
